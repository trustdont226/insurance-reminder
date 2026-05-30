"""
app.py  –  Insurance Renewal Reminder  (Flask entry point)

Application factory layer kept thin on purpose. All business logic lives in
the `core/` package.

Run:    python app.py
Open:   http://localhost:5000
"""

import io
import json
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS

from core import scheduler
from core.config import DEBUG, HOST, PORT
from core.database import (
    delete_record,
    get_all_records,
    get_record_by_id,
    get_setting,
    init_db,
    insert_record,
    set_setting,
    update_record,
)
from core.config import using_supabase
from core.pdf_processor import extract_insurance_data
from core.storage import (
    delete_temp,
    save_pdf,
    save_pdf_temp,
    upload_to_supabase,
)
from core.whatsapp import DEFAULT_TEMPLATE, build_reminder_message, build_whatsapp_url


# ── App setup ────────────────────────────────────────────────────────────────

app = Flask(__name__)
CORS(app)

init_db()


def get_reminder_days() -> int:
    """Read reminder days from DB (agent can change it from UI)."""
    return int(get_setting("reminder_days", "3"))


# ── Helpers ──────────────────────────────────────────────────────────────────

def _enrich(record: dict) -> dict:
    """Add `days_remaining` and `status` fields to a DB record."""
    record = dict(record)
    renewal_raw = record.get("renewal_date")
    record["days_remaining"] = None
    record["status"] = "unknown"

    if renewal_raw:
        try:
            renewal_dt = datetime.strptime(renewal_raw, "%d/%m/%Y").date()
            days = (renewal_dt - datetime.now().date()).days
            record["days_remaining"] = days

            if record.get("reminder_sent"):
                record["status"] = "sent"
            elif days < 0:
                record["status"] = "overdue"
            elif days <= get_reminder_days():
                record["status"] = "due_soon"
            elif days <= 30:
                record["status"] = "upcoming"
            else:
                record["status"] = "ok"
        except ValueError:
            pass

    return record


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload():
    """
    Accept one or more PDF files, extract insurance data, and persist:
      • OCR runs on a local temp copy
      • PDF is uploaded to Supabase Storage (if configured)
      • Temp copy is deleted afterwards
    """
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files provided"}), 400

    results = []
    use_supabase = using_supabase()

    for file in files:
        if not file.filename:
            continue
        if not file.filename.lower().endswith(".pdf"):
            results.append({"filename": file.filename, "success": False,
                             "error": "Only PDF files are accepted"})
            continue

        temp_path = None
        storage_key = None
        try:
            if use_supabase:
                temp_path = save_pdf_temp(file, file.filename)
                ocr_path  = temp_path
            else:
                # Local-only fallback
                ocr_path = save_pdf(file, file.filename)

            data = extract_insurance_data(ocr_path)
            data["pdf_filename"] = file.filename

            if use_supabase:
                try:
                    storage_key = upload_to_supabase(temp_path, file.filename)
                    data["pdf_storage_path"] = storage_key
                except Exception as upload_err:
                    # OCR succeeded — keep the data even if Supabase upload fails
                    print(f"[Storage] Upload failed for {file.filename}: {upload_err}")

            record_id = insert_record(data)
            data["id"] = record_id
            results.append({"filename": file.filename, "success": True, "data": data})

        except Exception as exc:
            results.append({"filename": file.filename, "success": False,
                             "error": str(exc)})
        finally:
            if temp_path:
                delete_temp(temp_path)

    return jsonify(results)


@app.route("/api/records", methods=["GET"])
def get_records():
    records = [_enrich(r) for r in get_all_records()]
    return jsonify(records)


@app.route("/api/records/<int:record_id>", methods=["PUT"])
def update(record_id):
    data = request.get_json(silent=True) or {}
    update_record(record_id, data)
    record = get_record_by_id(record_id)
    return jsonify(_enrich(record)) if record else (jsonify({"error": "Not found"}), 404)


@app.route("/api/records/<int:record_id>", methods=["DELETE"])
def delete(record_id):
    delete_record(record_id)
    return jsonify({"success": True})


@app.route("/api/send-reminder/<int:record_id>", methods=["POST"])
def send_one(record_id):
    """
    Build a Click-to-Chat WhatsApp URL with the pre-filled message.
    Frontend opens the URL in a new tab — agent clicks Send in WhatsApp.
    """
    record = get_record_by_id(record_id)
    if not record:
        return jsonify({"error": "Record not found"}), 404

    result = build_whatsapp_url(record)
    if result["success"]:
        update_record(record_id, {
            "reminder_sent": 1,
            "reminder_sent_at": datetime.now().isoformat(timespec="seconds"),
        })
    return jsonify(result)


@app.route("/api/records/<int:record_id>/mark-unsent", methods=["POST"])
def mark_unsent(record_id):
    """Undo the 'sent' flag for a record."""
    record = get_record_by_id(record_id)
    if not record:
        return jsonify({"error": "Record not found"}), 404
    update_record(record_id, {"reminder_sent": 0, "reminder_sent_at": None})
    return jsonify({"success": True})


@app.route("/api/send-reminders", methods=["POST"])
def send_bulk():
    """Build URLs for all records due within reminder window."""
    records = get_all_records()
    today = datetime.now().date()
    pending, skipped = [], []

    for rec in records:
        rec_id = rec["id"]

        if rec.get("reminder_sent"):
            skipped.append({"id": rec_id, "reason": "already sent"})
            continue

        if not rec.get("mobile_number"):
            skipped.append({"id": rec_id, "reason": "no mobile number"})
            continue

        try:
            renewal_dt = datetime.strptime(rec["renewal_date"], "%d/%m/%Y").date()
            days = (renewal_dt - today).days
        except (ValueError, TypeError):
            skipped.append({"id": rec_id, "reason": "invalid renewal date"})
            continue

        if not (0 <= days <= get_reminder_days()):
            skipped.append({"id": rec_id, "reason": f"{days} days away"})
            continue

        result = build_whatsapp_url(rec)
        if result["success"]:
            pending.append({
                "id":     rec_id,
                "owner":  rec.get("owner_name") or "Customer",
                "mobile": rec.get("mobile_number"),
                "url":    result["url"],
            })
        else:
            skipped.append({"id": rec_id, "reason": result.get("error", "unknown")})

    return jsonify({"pending": pending, "skipped": skipped})


@app.route("/api/export", methods=["GET"])
def export_excel():
    """Generate and return a formatted Excel file of all insurance records."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    records = get_all_records()
    today   = datetime.now().date()
    wb      = Workbook()
    ws      = wb.active
    ws.title = "Insurance Records"

    HEADER_BG   = "1A56DB"
    HEADER_FG   = "FFFFFF"
    ALT_ROW     = "F0F4FF"
    OVERDUE_BG  = "FEE2E2"
    DUE_BG      = "FEF3C7"
    SENT_BG     = "EDE9FE"
    OK_BG       = "D1FAE5"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = [
        ("Sr. No.",           10),
        ("Owner Name",        28),
        ("Mobile Number",     16),
        ("Vehicle No.",       16),
        ("Policy Number",     28),
        ("Chassis No.",       22),
        ("Engine No.",        20),
        ("Renewal Date",      16),
        ("Payment Due Date",  18),
        ("Premium (Rs.)",     16),
        ("Make / Model",      20),
        ("Issuing Office",    16),
        ("Status",            14),
        ("Reminder Sent",     16),
        ("Reminder Sent At",  20),
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    title_cell = ws["A1"]
    title_cell.value     = f"Insurance Renewal Records — Exported {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    title_cell.font      = Font(bold=True, size=13, color=HEADER_FG)
    title_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 22

    ws.freeze_panes = "A3"

    for row_idx, rec in enumerate(records, start=1):
        status = "—"
        row_bg = None
        try:
            renewal_dt = datetime.strptime(rec["renewal_date"], "%d/%m/%Y").date()
            days = (renewal_dt - today).days
            if rec.get("reminder_sent"):
                status = "Reminder Sent"
                row_bg = SENT_BG
            elif days < 0:
                status = f"Overdue ({abs(days)}d)"
                row_bg = OVERDUE_BG
            elif days <= get_reminder_days():
                status = f"Due in {days}d"
                row_bg = DUE_BG
            elif days <= 30:
                status = f"In {days} days"
            else:
                status = "OK"
                row_bg = OK_BG
        except (ValueError, TypeError):
            status = "Unknown"

        premium = rec.get("premium_amount") or ""
        try:
            premium = f"{int(float(premium)):,}"
        except (ValueError, TypeError):
            pass

        sent_at = rec.get("reminder_sent_at") or ""
        if sent_at:
            try:
                sent_at = datetime.fromisoformat(sent_at).strftime("%d/%m/%Y %H:%M")
            except Exception:
                pass

        row_data = [
            row_idx,
            rec.get("owner_name")         or "",
            rec.get("mobile_number")      or "",
            rec.get("vehicle_number")     or "",
            rec.get("policy_number")      or "",
            rec.get("chassis_number")     or "",
            rec.get("engine_number")      or "",
            rec.get("renewal_date")       or "",
            rec.get("payment_due_date")   or "",
            premium,
            rec.get("vehicle_make_model") or "",
            rec.get("issuing_office")     or "",
            status,
            "Yes" if rec.get("reminder_sent") else "No",
            sent_at,
        ]

        excel_row = row_idx + 2
        alt_bg    = ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
        fill_bg   = row_bg or alt_bg

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=fill_bg)
            cell.border    = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col_idx in (1, 13, 14) else "left")
            cell.font      = Font(size=9)
        ws.row_dimensions[excel_row].height = 18

    ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{len(records) + 2}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(bold=True, size=14, color=HEADER_FG)
    ws2["A1"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws2.merge_cells("A1:B1")

    overdue = due_soon = sent_count = ok_count = 0
    for rec in records:
        if rec.get("reminder_sent"):
            sent_count += 1
        try:
            days = (datetime.strptime(rec["renewal_date"], "%d/%m/%Y").date() - today).days
            if days < 0:                       overdue  += 1
            elif days <= get_reminder_days():  due_soon += 1
            else:                              ok_count += 1
        except (ValueError, TypeError):
            pass

    summary_data = [
        ("Total Records",    len(records)),
        ("Overdue",          overdue),
        ("Due Soon",         due_soon),
        ("Reminders Sent",   sent_count),
        ("OK",               ok_count),
        ("Export Date",      datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for i, (label, value) in enumerate(summary_data, start=2):
        ws2[f"A{i}"] = label
        ws2[f"B{i}"] = value
        ws2[f"A{i}"].font = Font(bold=True, size=10)
        ws2[f"B{i}"].font = Font(size=10)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"insurance_records_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/settings", methods=["GET"])
def get_settings():
    last_result_raw = get_setting("scheduler_last_result", "")
    try:
        last_result = json.loads(last_result_raw) if last_result_raw else {}
    except Exception:
        last_result = {}

    return jsonify({
        "reminder_days":         get_reminder_days(),
        "scheduler_enabled":     get_setting("scheduler_enabled", "false") == "true",
        "scheduler_time":        get_setting("scheduler_time", "09:00"),
        "scheduler_last_run":    get_setting("scheduler_last_run", ""),
        "scheduler_last_result": last_result,
        "scheduler_next_run":    scheduler.get_next_run_time(),
        "message_template":      get_setting("message_template", DEFAULT_TEMPLATE),
        "default_template":      DEFAULT_TEMPLATE,
    })


@app.route("/api/settings", methods=["POST"])
def save_settings():
    data = request.get_json(silent=True) or {}

    if "reminder_days" in data:
        try:
            days = int(data["reminder_days"])
            if not (1 <= days <= 60):
                raise ValueError
            set_setting("reminder_days", days)
        except ValueError:
            return jsonify({"error": "reminder_days must be 1–60"}), 400

    if "scheduler_enabled" in data:
        set_setting("scheduler_enabled", "true" if data["scheduler_enabled"] else "false")

    if "scheduler_time" in data:
        t = str(data["scheduler_time"])
        try:
            h, m = map(int, t.split(":"))
            if not (0 <= h <= 23 and 0 <= m <= 59):
                raise ValueError
            set_setting("scheduler_time", t)
        except ValueError:
            return jsonify({"error": "scheduler_time must be HH:MM"}), 400

    if "scheduler_enabled" in data or "scheduler_time" in data:
        scheduler.apply_settings()

    if "message_template" in data:
        template = str(data["message_template"]).strip()
        if not template:
            return jsonify({"error": "Message template cannot be empty"}), 400
        if len(template) > 4000:
            return jsonify({"error": "Message template too long (max 4000 chars)"}), 400
        set_setting("message_template", template)

    return jsonify({"success": True})


@app.route("/api/preview-message/<int:record_id>", methods=["POST"])
def preview_message(record_id):
    """Render the current template with a record's actual data."""
    record = get_record_by_id(record_id)
    if not record:
        return jsonify({"error": "Record not found"}), 404

    data = request.get_json(silent=True) or {}
    if "template" in data:
        original = get_setting("message_template", "")
        set_setting("message_template", data["template"])
        try:
            text = build_reminder_message(record)
        finally:
            set_setting("message_template", original)
    else:
        text = build_reminder_message(record)

    return jsonify({"success": True, "message": text})


@app.route("/api/scheduler/run-now", methods=["POST"])
def run_scheduler_now():
    """Manually trigger the scheduler job immediately."""
    try:
        scheduler.run_scheduled_reminders()
        last_result_raw = get_setting("scheduler_last_result", "")
        result = json.loads(last_result_raw) if last_result_raw else {}
        return jsonify({"success": True, "result": result})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/stats", methods=["GET"])
def stats():
    records = get_all_records()
    today = datetime.now().date()
    total = len(records)
    due_soon = 0
    reminder_sent = 0

    for rec in records:
        if rec.get("reminder_sent"):
            reminder_sent += 1
        try:
            renewal_dt = datetime.strptime(rec["renewal_date"], "%d/%m/%Y").date()
            days = (renewal_dt - today).days
            if 0 <= days <= get_reminder_days():
                due_soon += 1
        except (ValueError, TypeError):
            pass

    return jsonify({"total": total, "due_soon": due_soon, "reminder_sent": reminder_sent})


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n  Insurance Renewal Reminder App")
    print(f"  Open: http://localhost:{PORT}\n")

    # Pre-load EasyOCR models at startup so the first upload isn't slow.
    print("  Loading OCR engine (one-time setup)...")
    from core.pdf_processor import _get_reader
    _get_reader()
    print("  OCR engine ready.")

    # Start background scheduler
    scheduler.start()
    scheduler.apply_settings()
    print()

    app.run(debug=DEBUG, host=HOST, port=PORT, use_reloader=False)
